import json
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import re
import os
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def delete_html_files():
    """Delete all .html files in the current working directory."""
    deleted_files = 0
    current_dir = os.getcwd()
    
    for filename in os.listdir(current_dir):
        if filename.endswith(".html"):
            try:
                os.remove(filename)
                print(f"Deleted: {filename}")
                deleted_files += 1
            except Exception as e:
                print(f"Error deleting {filename}: {e}")
    
    print(f"\nTotal {deleted_files} HTML files deleted from {current_dir}")



def write_code(filename:str, code:str):
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(code)

def fetch_and_save_html(url: str, filename: str,code:bool=False,pretty:bool=True):
    try:
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--window-size=1920,1080")

        driver = webdriver.Chrome(options=chrome_options)
        try:
            driver.get(url)

            for _ in range(5):
                driver.refresh()
                time.sleep(1)
            
            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.flex.flex-col"))
            )
            
            html = driver.page_source
            if pretty==True:
                soup = BeautifulSoup(html, 'html.parser').prettify()
            if pretty==False:
                soup = BeautifulSoup(html, 'html.parser')

            if code==True:
                try:
                    write_code(filename, soup)
                except:
                    write_code(filename, str(soup))
            else:
                return soup
        finally:
            driver.quit()
            print(f"\n Got The html code from : {url}")
    except Exception as e:
        print(f"Error: {e}")
    return 'success'

def get_href(url:str,filename:str):
    hrefs=[]
    with open(filename, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')
        for a_tag in soup.find_all('a', href=True):
            href = a_tag['href']
            if href.startswith(url) and not href.endswith('/standings/'):
                hrefs.append(href)
    print(f"\n Got {len(hrefs)} links from the html file")
    return hrefs

def get_date(filename:str,row:str,output_file:str):
    with open(filename, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')
        script_tag = soup.find('script', type='application/ld+json')
        game_time = ''
        if script_tag:
            try:
                data = json.loads(script_tag.string)
                if isinstance(data, list):
                    for item in data:
                        if 'startDate' in item:
                            game_time = item['startDate']
                            break
                elif 'startDate' in data:
                    game_time = data['startDate'][:10]
            except Exception as e:
                print(f"Error parsing JSON-LD: {e}")
    print(game_time)
    wb = load_workbook(f'{output_file}')
    ws = wb.active
    ws[f'C{row}'] = game_time
    wb.save(f'{output_file}')

def get_teams(filename:str,row:str,output_file:str):
    with open(filename, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')
        script_tag = soup.find('script', type='application/ld+json')
        home_team = ''
        away_team = ''
        if script_tag:
            try:
                data = json.loads(script_tag.string)
                if isinstance(data, list):
                    for item in data:
                        if 'homeTeam' in item and 'awayTeam' in item:
                            home_team = item['homeTeam']['name']
                            away_team = item['awayTeam']['name']
                            break
                else:
                    if 'homeTeam' in data and 'awayTeam' in data:
                        home_team = data['homeTeam']['name']
                        away_team = data['awayTeam']['name']
            except Exception as e:
                print(f"Error parsing JSON-LD: {e}")
    print(f"Home Team: {home_team}, Away Team: {away_team}")
    wb = load_workbook(f'{output_file}')
    ws = wb.active
    ws[f'D{row}'] = home_team
    ws[f'E{row}'] = away_team
    wb.save(f'{output_file}')

def get_results(filename:str,row:str,output_file:str):
        ratios=None
        # Read the HTML file
        with open(f"{filename}", "r", encoding="utf-8") as file:
            html = file.read()

        # Look for the embedded JSON-like structure
        match = re.search(r'"text":"Final\\u0026nbsp;result.*?"', html)

        if match:
            raw_result = match.group(0)

            # Convert unicode escapes and HTML tags to readable format
            decoded = bytes(raw_result, "utf-8").decode("unicode_escape")
            cleaned = BeautifulSoup(decoded, "html.parser").get_text()
            cleaned = cleaned.replace("Final\u0026nbsp;result", "").strip()
            # Extract all numbers in a:b format
            scores = re.findall(r'(\d+\s*:\s*\d+)', cleaned)
            if scores:
                ratios = scores
            else:
                ratios = ["No score found"]
        else:
            ratios = []
            print("Match result not found.")
        print("Extracted ratios:", ratios)
        wb = load_workbook(f'{output_file}')
        ws = wb.active
        cells = [(f'F{row}', f'G{row}'), (f'H{row}', f'I{row}'), (f'J{row}', f'K{row}')]
        for idx, ratio in enumerate(ratios[:3]):
            parts = ratio.split(':')
            if len(parts) == 2 and idx < len(cells):
                ws[cells[idx][0]] = parts[0]
                ws[cells[idx][1]] = parts[1]
        wb.save(f'{output_file}')

def write_odds_to_excel(odds, columns, row, excel_path,purpose:str):
    wb = load_workbook(excel_path)
    ws = wb.active

    col_idx = 0


    # Always follow the order: 'Pinnacle', 'Bet365', '1xBet'
    for bookmaker in ['Pinnacle', 'Bet365', '1xBet']:
        if bookmaker in odds:
            if purpose == '1x2':
                ws[f'{columns[col_idx]}{row}'] = odds[bookmaker].get('1', '')
                ws[f'{columns[col_idx + 1]}{row}'] = odds[bookmaker].get('X', '')
                ws[f'{columns[col_idx + 2]}{row}'] = odds[bookmaker].get('2', '')
                col_idx += 3
            elif purpose == 'over/under':
                ws[f'{columns[col_idx]}{row}'] = odds[bookmaker].get('Over', '')
                ws[f'{columns[col_idx + 1]}{row}'] = odds[bookmaker].get('Under', '')
                col_idx += 2
            elif purpose == 'yes/no':
                ws[f'{columns[col_idx]}{row}'] = odds[bookmaker].get('Yes', '')
                ws[f'{columns[col_idx + 1]}{row}'] = odds[bookmaker].get('No', '')
                col_idx += 2


    wb.save(excel_path)
    print(odds[bookmaker])
    print(f"✅ Odds written to {excel_path}")

def extract_odds_by_bookmaker(filename, row_selector, odds_count, odds_labels):
    with open(filename, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'html.parser')

    odds_data = {}
    target_bookmakers = {
        "1xbet": "1xBet",
        "pinnacle": "Pinnacle",
        "bet365": "Bet365"
    }

    rows = soup.find_all('div', {'data-testid': row_selector})
    for row in rows:
        name_tag = row.find('p', {'data-testid': 'outrights-expanded-bookmaker-name'})
        if not name_tag:
            continue
        name = name_tag.get_text(strip=True).lower()
        if name in target_bookmakers:
            odds_tags = row.find_all('div', {'data-testid': 'odd-container'})
            if len(odds_tags) >= odds_count:
                odds = [odds_tags[i].get_text(strip=True).replace(",", ".") for i in range(odds_count)]
                odds_data[target_bookmakers[name]] = dict(zip(odds_labels, odds))
            else:
                print(f"[!] Not enough odds found for {name}")
    for bm in target_bookmakers.values():
        if bm not in odds_data:
            print(f"[!] Odds not found for: {bm}")
    return odds_data

def get_1x2_odds_by_bookmaker(filename):
    return extract_odds_by_bookmaker(
        filename,
        row_selector='over-under-expanded-row',
        odds_count=3,
        odds_labels=['1', 'X', '2']
    )

def get_over_under_odds_by_bookmaker(filename):
    return extract_odds_by_bookmaker(
        filename,
        row_selector='over-under-expanded-row',
        odds_count=2,
        odds_labels=['Over', 'Under']
    )

def get_yes_no_odds_by_bookmaker(filename):
    return extract_odds_by_bookmaker(
        filename,
        row_selector='over-under-expanded-row',
        odds_count=2,
        odds_labels=['Yes', 'No']
    )











if __name__ == "__main__":
    url = input("\nEnter the url\nHINT: https://www.oddsportal.com/football/LEAGUE-COUNTRY/LEAGUE-NAME/standings\n-2024-2025 after league-name if needed\n : ")
    # url = "https://www.oddsportal.com/football/spain/laliga/standings"
    league_country = url.split("/football/")[-1].split("/")[0]
    league_name = url.split(f"{league_country}/")[-1].split("/")[0]
    output_file = str(input("\nEnter Output Excel File Name:"))
    output_file = output_file if output_file.endswith('.xlsx') else output_file + '.xlsx'
    print(f"League Country: {league_country}")
    print(f"League Name: {league_name}")

    fetch_and_save_html(url, f"{league_country}-{league_name}.html",True)
    

    links=get_href(url.split("/standing")[0], f"{league_country}-{league_name}.html")
    links = list(set(links))
    counter=0
    for link in links:
        counter+=1
        row= counter+8
        if row > 13:
            try:
                print(f"\n{link}")

                # Sanitize filename to avoid leading slashes and invalid characters
                raw_filename = link.split(league_name)[-1]
                sanitized_filename = raw_filename.lstrip('/').replace('/', '_').replace('\\', '_') + ".html"
                # output_file = f"{league_country}-{league_name}.xlsx"
                

                fetch_and_save_html(link, sanitized_filename,True)

                get_date(sanitized_filename, str(row), output_file)

                get_teams(sanitized_filename, str(row), output_file)

                get_results(sanitized_filename, str(row), output_file)
                
                print("\nWaiting to fetch odds for FULL TIME 1x2")
                for _ in range(2):
                    fetch = fetch_and_save_html(link+"#1X2;2", sanitized_filename, True,  )
                    if fetch == 'success':
                        odds = get_1x2_odds_by_bookmaker(sanitized_filename)
                        # Check if any value in odds dict is empty
                        columns = ['M','N','O','P','Q','R','S','T','U']
                        if any(vv for v in odds.values() for vv in v.values()):
                            write_odds_to_excel(odds, columns, row, output_file,'1x2')
                            break
                    
                
                print("\nWaiting to fetch odds for 1st HALF 1x2")
                for _ in range(2):
                    fetch = fetch_and_save_html(link+"#1X2;3", sanitized_filename, True, )
                    if fetch == 'success':
                        odds = get_1x2_odds_by_bookmaker(sanitized_filename)
                        columns = ['W','X','Y','Z','AA','AB','AC','AD','AE']
                        if any(vv for v in odds.values() for vv in v.values()):
                            write_odds_to_excel(odds, columns, row, output_file,'1x2')
                            break



                # OVER/UNDER FULL TIME
                print("\nWaiting to fetch odds for FULL TIME 1.5 over/under")
                for _ in range(2):
                    fetch = fetch_and_save_html(link+"#over-under;2;1.50;0", sanitized_filename, True, )
                    if fetch == 'success':
                        odds = get_over_under_odds_by_bookmaker(sanitized_filename)
                        columns = ['AG','AH','AI','AJ','AK','AL']
                        if any(vv for v in odds.values() for vv in v.values()):
                            
                            write_odds_to_excel(odds, columns, row, output_file,'over/under')
                            break

                print("\nWaiting to fetch odds for FULL TIME 2.5 over/under")
                for _ in range(2):
                    fetch = fetch_and_save_html(link+"#over-under;2;2.50;0", sanitized_filename, True, )
                    if fetch == 'success':
                        odds = get_over_under_odds_by_bookmaker(sanitized_filename)
                        columns = ['AN','AO','AP','AQ','AR','AS']
                        if any(vv for v in odds.values() for vv in v.values()):
                            
                            write_odds_to_excel(odds, columns, row, output_file,'over/under')
                            break

                print("\nWaiting to fetch odds for FULL TIME 3.5 over/under")
                for _ in range(2):
                    fetch = fetch_and_save_html(link+"#over-under;2;3.50;0", sanitized_filename, True, )
                    if fetch == 'success':
                        odds = get_over_under_odds_by_bookmaker(sanitized_filename)
                        columns = ['AU','AV','AW','AX','AY','AZ']
                        if any(vv for v in odds.values() for vv in v.values()):
                            
                            write_odds_to_excel(odds, columns, row, output_file,'over/under')
                            break

                

                # OVER/UNDER 1ST HALF
                print("\nWaiting to fetch odds for 1ST HALF 0.5 over/under")
                for _ in range(2):
                    fetch = fetch_and_save_html(link+"#over-under;3;0.50;0", sanitized_filename, True, )
                    if fetch == 'success':
                        odds = get_over_under_odds_by_bookmaker(sanitized_filename)
                        columns = ['BB','BC','BD','BE','BF','BG']
                        if any(vv for v in odds.values() for vv in v.values()):
                            
                            write_odds_to_excel(odds, columns, row, output_file,'over/under')
                            break

                print("\nWaiting to fetch odds for 1ST HALF 1.5 over/under")
                for _ in range(2):
                    fetch = fetch_and_save_html(link+"#over-under;3;1.50;0", sanitized_filename, True, )
                    if fetch == 'success':
                        odds = get_over_under_odds_by_bookmaker(sanitized_filename)
                        columns = ['BI','BJ','BK','BL','BM','BN']
                        if any(vv for v in odds.values() for vv in v.values()):
                            
                            write_odds_to_excel(odds, columns, row, output_file,'over/under')
                            break

                print("\nWaiting to fetch odds for 1ST HALF 2.5 over/under")
                for _ in range(2):
                    fetch = fetch_and_save_html(link+"#over-under;3;2.50;0", sanitized_filename, True, )
                    if fetch == 'success':
                        odds = get_over_under_odds_by_bookmaker(sanitized_filename)
                        columns = ['BP','BQ','BR','BS','BT','BU']
                        if any(vv for v in odds.values() for vv in v.values()):
                            
                            write_odds_to_excel(odds, columns, row, output_file,'over/under')
                            break
            


                print("\nWaiting to fetch odds for BOTH TEAMS TO SCORE")
                for _ in range(2):
                    fetch = fetch_and_save_html(link+"#bts;2", sanitized_filename, True, )
                    if fetch == 'success':
                        odds = get_yes_no_odds_by_bookmaker(sanitized_filename)
                        columns = ['BW','BX','BY','BZ','CA','CB']
                        if any(vv for v in odds.values() for vv in v.values()):
                            write_odds_to_excel(odds, columns, row, output_file,'yes/no')
                            break
            except:
                pass

            # delete_html_files()

input("\n\n\nPress Enter to Exit . . . ")