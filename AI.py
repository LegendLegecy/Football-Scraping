import json
import time
import re
import os
import random
from urllib.parse import urlparse
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from concurrent.futures import ThreadPoolExecutor
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

start_time = time.time()

def get_json_data(key_name: str, config_path='config.json'):
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    return config.get(key_name)

def ensure_excel_file(filepath):
    if not os.path.exists(filepath):
        wb = Workbook()
        wb.save(filepath)
        print(f"[📄] Created new Excel file: {filepath}", flush=True)

def write_code(filename: str, code: str):
    try:
        if "<html" not in code.lower():
            print(f"⚠️ HTML seems invalid for {filename}, but saving anyway", flush=True)
            with open("suspect_pages.log", "a") as log:
                log.write(f"{filename}\n")
        path = os.path.abspath(filename)
        print(f"[💾] Saving to: {path}", flush=True)
        with open(path, 'w', encoding='utf-8') as f:
            f.write(code)
    except Exception as e:
        print(f"[ERROR] Could not write to {filename}: {e}", flush=True)

def sanitize_filename(url, league_name):
    path = urlparse(url).path
    tail = path.split(f"/{league_name}-2024-2025/")[-1]
    if not tail.strip():
        tail = "index"
    return f"{league_name}-{tail.replace('/', '_').replace('\\', '_')}.html"

def create_driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--disable-dev-shm-usage")
    return webdriver.Chrome(options=chrome_options)

def fetch_and_save_html(driver, url: str, filename: str, code: bool = False, pretty: bool = False):
    try:
        driver.get(url)
        # Wait for main content to load
        for _ in range(10):
            driver.refresh()
            time.sleep(2)
        time.sleep(30)
        
        # Scroll to load dynamic content
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        if pretty:
            soup = soup.prettify()

        if code:
            try:
                write_code(filename, soup)
            except:
                write_code(filename, str(soup))
        else:
            return soup

        print(f"\n✅ Got the HTML code from: {url}", flush=True)

    except Exception as e:
        print(f"❌ Error fetching HTML from {url}: {e}", flush=True)

def get_href(url: str, filename: str):
    hrefs = []
    with open(filename, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')
        for a_tag in soup.find_all('a', href=True):
            href = a_tag['href']
            if href.startswith(url) and not href.endswith('/standings/'):
                hrefs.append(href)
    print(f"\n Got {len(hrefs)} links from the html file", flush=True)
    return hrefs

def get_date(filename: str, row: str, output_file: str):
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            soup = BeautifulSoup(f, 'html.parser')
            script_tag = soup.find('script', type='application/ld+json')
            game_time = ''
            if script_tag:
                data = json.loads(script_tag.string)
                if isinstance(data, list):
                    for item in data:
                        if 'startDate' in item:
                            game_time = item['startDate']
                            break
                elif 'startDate' in data:
                    game_time = data['startDate'][:10]
        wb = load_workbook(output_file)
        ws = wb.active
        ws[f'C{row}'] = game_time
        wb.save(output_file)
    except Exception as e:
        print(f"❌ Failed to extract date from {filename}: {e}", flush=True)

def get_teams(filename: str, row: str, output_file: str):
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            soup = BeautifulSoup(f, 'html.parser')
            script_tag = soup.find('script', type='application/ld+json')
            home_team = ''
            away_team = ''
            if script_tag:
                data = json.loads(script_tag.string)
                if isinstance(data, list):
                    for item in data:
                        if 'homeTeam' in item and 'awayTeam' in item:
                            home_team = item['homeTeam']['name']
                            away_team = item['awayTeam']['name']
                            break
                else:
                    if 'homeTeam' in data and 'awayTeam' in data:
                        home_team = data['homeTeam']['name']
                        away_team = data['awayTeam']['name']
        wb = load_workbook(output_file)
        ws = wb.active
        ws[f'D{row}'] = home_team
        ws[f'E{row}'] = away_team
        wb.save(output_file)
    except Exception as e:
        print(f"❌ Failed to extract teams from {filename}: {e}", flush=True)

def get_results(filename: str, row: str, output_file: str):
    try:
        with open(filename, "r", encoding="utf-8") as file:
            html = file.read()
        match = re.search(r'"text":"Final\\u0026nbsp;result.*?"', html, re.UNICODE)

        ratios = []
        if match:
            raw_result = match.group(0)
            decoded = bytes(raw_result, "utf-8").decode("unicode_escape")
            cleaned = BeautifulSoup(decoded, "html.parser").get_text()
            cleaned = cleaned.replace("Final&nbsp;result", "").strip()
            scores = re.findall(r'(\d+\s*:\s*\d+)', cleaned)
            if scores:
                ratios = scores
            else:
                ratios = ["No score found"]
        wb = load_workbook(output_file)
        ws = wb.active
        cells = [(f'F{row}', f'G{row}'), (f'H{row}', f'I{row}'), (f'J{row}', f'K{row}')]
        for idx, ratio in enumerate(ratios[:3]):
            parts = ratio.split(':')
            if len(parts) == 2 and idx < len(cells):
                ws[cells[idx][0]] = parts[0]
                ws[cells[idx][1]] = parts[1]
        wb.save(output_file)
    except Exception as e:
        print(f"❌ Failed to extract results from {filename}: {e}", flush=True)

def extract_odds_by_bookmaker(filename, row_selector, odds_count, odds_labels):
    with open(filename, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'html.parser')
    odds_data = {}
    target_bookmakers = {"1xbet": "1xBet", "pinnacle": "Pinnacle", "bet365": "Bet365"}
    
    # Try both selectors as the site might change them
    rows = soup.find_all('div', {'class': 'flex border-b border-black-borders min-h-[38px]'})
    if not rows:
        rows = soup.find_all('div', {'data-testid': row_selector})
    
    for row in rows:
        name_tag = row.find('p', class_='text-xs font-normal text-gray-dark')
        if not name_tag:
            name_tag = row.find('p', {'data-testid': 'outrights-expanded-bookmaker-name'})
        if not name_tag:
            continue
            
        name = name_tag.get_text(strip=True).lower()
        if name in target_bookmakers:
            odds_tags = row.find_all('div', class_='flex h-[38px] items-center justify-center border-r border-black-borders')
            if not odds_tags:
                odds_tags = row.find_all('div', {'data-testid': 'odd-container'})
                
            if len(odds_tags) >= odds_count:
                odds = [odds_tags[i].get_text(strip=True).replace(",", ".") for i in range(odds_count)]
                odds_data[target_bookmakers[name]] = dict(zip(odds_labels, odds))
    
    return odds_data

def write_odds_to_excel(odds, columns, row, excel_path, purpose: str):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        col_idx = 0
        for bookmaker in ['Pinnacle', 'Bet365', '1xBet']:
            if bookmaker in odds:
                if purpose == '1x2':
                    ws[f'{columns[col_idx]}{row}'] = odds[bookmaker].get('1', '')
                    ws[f'{columns[col_idx + 1]}{row}'] = odds[bookmaker].get('X', '')
                    ws[f'{columns[col_idx + 2]}{row}'] = odds[bookmaker].get('2', '')
                    col_idx += 3
                elif purpose == 'over/under':
                    ws[f'{columns[col_idx]}{row}'] = odds[bookmaker].get('Over', '')
                    ws[f'{columns[col_idx + 1]}{row}'] = odds[bookmaker].get('Under', '')
                    col_idx += 2
                elif purpose == 'yes/no':
                    ws[f'{columns[col_idx]}{row}'] = odds[bookmaker].get('Yes', '')
                    ws[f'{columns[col_idx + 1]}{row}'] = odds[bookmaker].get('No', '')
                    col_idx += 2

        wb.save(excel_path)
        print(f"✅ Odds written to {excel_path} [row {row} for {purpose}]", flush=True)
    except Exception as e:
        print(f"❌ Failed to write odds to Excel: {e}", flush=True)




def process_match(driver, link, row, league_country, league_name, output_file):
    try:
        print(f"\nProcessing match {row-8}: {link}")
        ensure_excel_file(output_file)
        sanitized_filename = sanitize_filename(link, league_name)
        
        # Get match page
        fetch_and_save_html(driver, link, sanitized_filename, code=True, pretty=False)
        get_date(sanitized_filename, str(row), output_file)
        get_teams(sanitized_filename, str(row), output_file)
        get_results(sanitized_filename, str(row), output_file)

        
    
        # Flat config list with function references
        odds_configs = [
            # 1X2
            ('#1x2;2', get_1x2_odds_by_bookmaker, ['M','N','O','P','Q','R','S','T','U'], '1x2'),
            ('#1x2;3', get_1x2_odds_by_bookmaker, ['W','X','Y','Z','AA','AB','AC','AD','AE'], '1x2'),
            ('#1x2;24', get_1x2_odds_by_bookmaker, ['AG','AH','AI','AJ','AK','AL','AM','AN','AO'], '1x2'),

            # Over/Under
            ('#over-under;2;1.50;0', get_over_under_odds_by_bookmaker, ['AQ','AR','AS','AT','AU','AV'], 'over/under'),
            ('#over-under;2;2.50;0', get_over_under_odds_by_bookmaker, ['AX','AY','AZ','BA','BB','BC'], 'over/under'),
            ('#over-under;2;3.50;0', get_over_under_odds_by_bookmaker, ['BE','BF','BG','BH','BI','BJ'], 'over/under'),
            ('#over-under;3;0.50;0', get_over_under_odds_by_bookmaker, ['BL','BM','BN','BO','BP','BQ'], 'over/under'),
            ('#over-under;3;1.50;0', get_over_under_odds_by_bookmaker, ['BS','BT','BU','BV','BW','BX'], 'over/under'),
            ('#over-under;3;2.50;0', get_over_under_odds_by_bookmaker, ['BZ','CA','CB','CC','CD','CE'], 'over/under'),
            ('#over-under;4;0.50;0', get_over_under_odds_by_bookmaker, ['CG','CH','CI','CJ','CK','CL'], 'over/under'),
            ('#over-under;4;1.50;0', get_over_under_odds_by_bookmaker, ['CN','CO','CP','CQ','CR','CS'], 'over/under'),
            ('#over-under;4;2.50;0', get_over_under_odds_by_bookmaker, ['CU','CV','CW','CX','CY','CZ'], 'over/under'),

            # Both Teams to Score
            ('#bts;2', get_yes_no_odds_by_bookmaker, ['DB','DC','DD','DE','DF','DG'], 'yes/no'),
        ]

        for suffix, odds_func, columns, purpose in odds_configs:
            try:
                fetch_and_save_html(driver, link + suffix, sanitized_filename, code=True, pretty=False)
                odds = odds_func(sanitized_filename)
                print()
                print()
                print()
                print()
                print()
                print(odds)
                print()
                print()
                print()
                print()
                print()
                if odds:  # Only write if we got odds
                    write_odds_to_excel(odds, columns, row, output_file, purpose)
                else:
                    print(f"⚠️ No odds found for {purpose} at {link + suffix}")
            except Exception as e:
                print(f"❌ Error processing {purpose} odds: {e}")

    except Exception as e:
        print(f"❌ Error in process_match for {link}: {e}", flush=True)




def run_parallel_scraping(links, league_country, league_name, output_file, max_workers=3):
    def process_link(link_row):
        link, row = link_row
        driver = create_driver()
        try:
            process_match(driver, link, row, league_country, league_name, output_file)
        finally:
            driver.quit()

    link_row_pairs = [(link, idx + 9) for idx, link in enumerate(links)]
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        executor.map(process_link, link_row_pairs)

if __name__ == "__main__":
    league_country = get_json_data('league-country').lower()
    league_name = get_json_data('league-name').lower().replace(" ", "-")
    output_file = f"{league_country}-{league_name}.xlsx"
    standings_file = f"{league_country}-{league_name}.html"
    
    # Initialize driver for standings page
    driver = create_driver()
    try:
        ensure_excel_file(output_file)
        fetch_and_save_html(driver,
            f"https://www.oddsportal.com/football/{league_country}/{league_name}-2024-2025/standings/",
            standings_file,
            code=True
        )
        links = get_href(f"https://www.oddsportal.com/football/{league_country}/{league_name}-2024-2025/", standings_file)
    finally:
        driver.quit()
    
    # Process matches in parallel
    run_parallel_scraping(links[:3], league_country, league_name, output_file)
    
    print(f"\n✅ Finished in {time.time() - start_time:.2f} seconds", flush=True)