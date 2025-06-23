import json
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import re
import os

start_time = time.time()

def get_json_data(key_name: str, config_path='config.json'):
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    return config.get(key_name)

def initialize_driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-notifications")
    return webdriver.Chrome(options=chrome_options)

def fetch_html(driver, url, max_retries=3):
    for attempt in range(max_retries):
        try:
            driver.get(url)
            # Wait for main content to load
            for _ in range(10):
                driver.refresh()
                time.sleep(2)
            time.sleep(30)
            
            return driver.page_source
        except Exception as e:
            print(f"Attempt {attempt + 1} failed: {e}")
            if attempt == max_retries - 1:
                raise
            time.sleep(2)

def save_html(content, filename):
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(content)

def get_hrefs_from_html(filename, base_url):
    with open(filename, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')
    return [a['href'] for a in soup.find_all('a', href=True) 
            if a['href'].startswith(base_url) and not a['href'].endswith('/standings/')]

def extract_json_ld_data(soup):
    script_tag = soup.find('script', type='application/ld+json')
    if not script_tag:
        return {}
    
    try:
        data = json.loads(script_tag.string)
        if isinstance(data, list):
            return data[0] if data else {}
        return data
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON-LD: {e}")
        return {}

def get_game_data(filename):
    with open(filename, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')
    
    data = extract_json_ld_data(soup)
    game_data = {
        'date': '',
        'home_team': '',
        'away_team': '',
        'results': []
    }
    
    if 'startDate' in data:
        game_data['date'] = data['startDate'][:10] if isinstance(data['startDate'], str) else ''
    
    if 'homeTeam' in data and 'awayTeam' in data:
        game_data['home_team'] = data['homeTeam'].get('name', '')
        game_data['away_team'] = data['awayTeam'].get('name', '')
    
    # Extract results
    html = str(soup)
    match = re.search(r'"text":"Final\\u0026nbsp;result.*?"', html)
    if match:
        raw_result = match.group(0)
        decoded = bytes(raw_result, "utf-8").decode("unicode_escape")
        cleaned = BeautifulSoup(decoded, "html.parser").get_text()
        cleaned = cleaned.replace("Final\u0026nbsp;result", "").strip()
        game_data['results'] = re.findall(r'(\d+\s*:\s*\d+)', cleaned)
    
    return game_data

def update_excel(output_file, row, game_data, odds_data):
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Basic game info
    ws[f'C{row}'] = game_data['date']
    ws[f'D{row}'] = game_data['home_team']
    ws[f'E{row}'] = game_data['away_team']
    
    # Results
    cells = [(f'F{row}', f'G{row}'), (f'H{row}', f'I{row}'), (f'J{row}', f'K{row}')]
    for idx, ratio in enumerate(game_data['results'][:3]):
        parts = ratio.split(':')
        if len(parts) == 2 and idx < len(cells):
            ws[cells[idx][0]] = parts[0]
            ws[cells[idx][1]] = parts[1]
    
    # Odds
    for purpose, columns in odds_data.get('columns', {}).items():
        for bookmaker in ['Pinnacle', 'Bet365', '1xBet']:
            if bookmaker in odds_data.get(purpose, {}):
                odds = odds_data[purpose][bookmaker]
                print()
                print()
                print()
                print()
                print()
                print(odds)
                print()
                print()
                print()
                print()
                print()
                if purpose == '1x2':
                    ws[f'{columns[0]}{row}'] = odds.get('1', '')
                    ws[f'{columns[1]}{row}'] = odds.get('X', '')
                    ws[f'{columns[2]}{row}'] = odds.get('2', '')
                elif purpose in ['over/under', 'yes/no']:
                    ws[f'{columns[0]}{row}'] = odds.get('Over', '') or odds.get('Yes', '')
                    ws[f'{columns[1]}{row}'] = odds.get('Under', '') or odds.get('No', '')
    
    wb.save(output_file)

def extract_odds(soup, row_selector, odds_count, odds_labels):
    odds_data = {}
    target_bookmakers = {
        "1xbet": "1xBet",
        "pinnacle": "Pinnacle",
        "bet365": "Bet365"
    }

    rows = soup.find_all('div', {'data-testid': row_selector})
    for row in rows:
        name_tag = row.find('p', {'data-testid': 'outrights-expanded-bookmaker-name'})
        if not name_tag:
            continue
        name = name_tag.get_text(strip=True).lower()
        if name in target_bookmakers:
            odds_tags = row.find_all('div', {'data-testid': 'odd-container'})
            if len(odds_tags) >= odds_count:
                odds = [odds_tags[i].get_text(strip=True).replace(",", ".") for i in range(odds_count)]
                odds_data[target_bookmakers[name]] = dict(zip(odds_labels, odds))
    
    return odds_data

def get_all_odds(driver, base_url):
    odds_types = {
        '1x2':[

            ('#1x2;2', get_1x2_odds_by_bookmaker, ['M','N','O','P','Q','R','S','T','U'], '1x2'),
            ('#1x2;3', get_1x2_odds_by_bookmaker, ['W','X','Y','Z','AA','AB','AC','AD','AE'], '1x2'),
            ('#1x2;24', get_1x2_odds_by_bookmaker, ['AG','AH','AI','AJ','AK','AL','AM','AN','AO'], '1x2'),
        ],
        'over/under':[

            # Over/Under
            ('#over-under;2;1.50;0', get_over_under_odds_by_bookmaker, ['AQ','AR','AS','AT','AU','AV'], 'over/under'),
            ('#over-under;2;2.50;0', get_over_under_odds_by_bookmaker, ['AX','AY','AZ','BA','BB','BC'], 'over/under'),
            ('#over-under;2;3.50;0', get_over_under_odds_by_bookmaker, ['BE','BF','BG','BH','BI','BJ'], 'over/under'),
            ('#over-under;3;0.50;0', get_over_under_odds_by_bookmaker, ['BL','BM','BN','BO','BP','BQ'], 'over/under'),
            ('#over-under;3;1.50;0', get_over_under_odds_by_bookmaker, ['BS','BT','BU','BV','BW','BX'], 'over/under'),
            ('#over-under;3;2.50;0', get_over_under_odds_by_bookmaker, ['BZ','CA','CB','CC','CD','CE'], 'over/under'),
            ('#over-under;4;0.50;0', get_over_under_odds_by_bookmaker, ['CG','CH','CI','CJ','CK','CL'], 'over/under'),
            ('#over-under;4;1.50;0', get_over_under_odds_by_bookmaker, ['CN','CO','CP','CQ','CR','CS'], 'over/under'),
            ('#over-under;4;2.50;0', get_over_under_odds_by_bookmaker, ['CU','CV','CW','CX','CY','CZ'], 'over/under'),

        ],
        'yes/no':[

            # Both Teams to Score
            ('#bts;2', get_yes_no_odds_by_bookmaker, ['DB','DC','DD','DE','DF','DG'], 'yes/no'),
        ]


        
    }
    
    all_odds = {'columns': {}}
    
    for purpose, urls in odds_types.items():
        all_odds[purpose] = {}
        for url_suffix, columns in urls:
            url = f"{base_url}{url_suffix}"
            html = fetch_html(driver, url)
            soup = BeautifulSoup(html, 'html.parser')
            
            if purpose == '1x2':
                odds = extract_odds(soup, 'over-under-expanded-row', 3, ['1', 'X', '2'])
            else:
                odds = extract_odds(soup, 'over-under-expanded-row', 2, 
                                   ['Over', 'Under'] if purpose == 'over/under' else ['Yes', 'No'])
            
            for bookmaker, values in odds.items():
                if bookmaker not in all_odds[purpose]:
                    all_odds[purpose][bookmaker] = {}
                all_odds[purpose][bookmaker].update(values)
            
            all_odds['columns'][purpose] = columns
    
    return all_odds

def main():
    league_country = get_json_data('league-country').lower()
    league_name = get_json_data('league-name').lower().replace(" ", "-")
    print(f"League Country: {league_country}")
    print(f"League Name: {league_name}")

    base_url = f"https://www.oddsportal.com/football/{league_country}/{league_name}-2024-2025"
    standings_url = f"{base_url}/standings/"
    output_file = f"{league_country}-{league_name}.xlsx"
    
    # Initialize driver once
    driver = initialize_driver()
    
    try:
        # Get standings page
        standings_html = fetch_html(driver, standings_url)
        standings_file = f"{league_country}-{league_name}.html"
        save_html(standings_html, standings_file)
        
        # Get match links
        links = get_hrefs_from_html(standings_file, base_url)
        
        for counter, link in enumerate(links[:1]):  # Process first match only for testing
            row = counter + 9
            print(f"\nProcessing match {counter + 1}: {link}")
            
            # Get match page
            match_html = fetch_html(driver, link)
            match_file = link.split(league_name)[-1].lstrip('/').replace('/', '_') + ".html"
            save_html(match_html, match_file)
            
            # Extract game data
            game_data = get_game_data(match_file)
            print(f"Date: {game_data['date']}")
            print(f"Teams: {game_data['home_team']} vs {game_data['away_team']}")
            print(f"Results: {game_data['results']}")
            
            # Get all odds in one go
            odds_data = get_all_odds(driver, link)
            
            # Update Excel
            update_excel(output_file, row, game_data, odds_data)
            print(f"✅ Data written to row {row}")
            
    finally:
        driver.quit()
    
    end_time = time.time()
    execution_time = end_time - start_time
    print(f"\nTotal execution time: {execution_time:.2f} seconds")
    
    with open("time.txt", 'w') as file:
        file.write(str(execution_time))

if __name__ == "__main__":
    main()